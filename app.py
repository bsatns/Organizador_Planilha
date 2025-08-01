from flask import Flask, render_template, request, send_file
import pandas as pd
import os
import uuid
from werkzeug.utils import secure_filename
import zipfile
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import numpy as np

app = Flask(__name__)

# Aumentar o limite de upload (exemplo: 100 MB)
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
RESULT_FOLDER = os.path.abspath(os.path.join(BASE_DIR, '..', 'results'))

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULT_FOLDER, exist_ok=True)

# Lista de colaboradores por setor
colaboradores_por_setor = {
    "Callcenter": ["Bruna", "Bruna Aguiar", "Beatriz", "Camila","Cláudia", "Edineide","Erica", "Maria Nayara", "Mayara", "Naara", "Tamires", "Raissa"],
    "Matriz": ["Marta", "Jamile"],
    "Cajazeiras": ["Ana Paula", "Agatha"]
}

def aplicar_formatacao_condicional(ws, coluna_obs):
    fill_cores = {
        "ATIVO WHATSAPP": "C6EFCE",
        "SEM INTERESSE": "FFC7CE",
        "VENDA": "00B050",
        "TEL NÃO ATENDEU": "FFEB9C",
        "SEM POSSIBILIDADES": "FF0000",
        "SEM CONTATO": "BDB76B",
        "SEM WHATSAPP": "8B008B",
        "SENDO TRAB": "FFA500"
    }

    inicio = 2
    fim = ws.max_row
    intervalo = f"{coluna_obs}{inicio}:{coluna_obs}{fim}"

    for status, cor in fill_cores.items():
        formula = f'${coluna_obs}{inicio}="{status}"'
        rule = FormulaRule(formula=[formula],
                           fill=PatternFill(start_color=cor, end_color=cor, fill_type="solid"))
        ws.conditional_formatting.add(intervalo, rule)

@app.route('/', methods=['GET'])
def index():
    return render_template("index.html", colaboradores=colaboradores_por_setor)

@app.route('/gerar_planilhas', methods=['POST'])
def gerar_planilhas():
    try:
        arquivos = request.files.getlist('Planilhas')
        if not arquivos:
            return "Nenhum arquivo enviado.", 400

        selecionados = request.form.getlist("colaboradores")
        if not selecionados:
            return "Nenhum colaborador selecionado.", 400

        seletor = request.form.get('SELETOR') == "sim"
        print ("Seletor", seletor)

        zip_name = f"planilhas_{uuid.uuid4().hex}.zip"
        zip_path = os.path.join(RESULT_FOLDER, zip_name)

        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file in arquivos:
                filename = secure_filename(file.filename)
                original_path = os.path.join(UPLOAD_FOLDER, filename)
                file.save(original_path)

                df = pd.read_excel(original_path)
                divisao = np.array_split(df, len(selecionados))

                for i, nome in enumerate(selecionados):
                    part_df = divisao[i] if i < len(divisao) else pd.DataFrame()

                    part_name = f"{nome}_{filename.rsplit('.', 1)[0]}.xlsx"
                    part_path = os.path.join(UPLOAD_FOLDER, part_name)
                    part_df.to_excel(part_path, index=False)

                    print ('Seletor', seletor)
                    if seletor:
                        wb = openpyxl.load_workbook(part_path)
                        ws = wb.active

                        obs_col_index = ws.max_column+1
                        obs_col_letter = get_column_letter(obs_col_index)
                        ws.cell(row=1, column=obs_col_index, value="OBSERVAÇÃO")

                        opcoes = [
                            "ATIVO WHATSAPP", "SEM INTERESSE", "VENDA", "TEL NÃO ATENDEU",
                            "SEM POSSIBILIDADES", "SEM CONTATO", "SEM WHATSAPP", "SENDO TRAB"
                        ]
                        dv = DataValidation(type="list", formula1='"' + ",".join(opcoes) + '"')
                        ws.add_data_validation(dv)
                        dv.add(f"{obs_col_letter}2:{obs_col_letter}{ws.max_row}")

                        for row in range(2, ws.max_row + 1):
                            ws[f'{obs_col_letter}{row}'] = ""

                        aplicar_formatacao_condicional(ws,obs_col_letter)

                        wb.save(part_path)

                    zipf.write(part_path, os.path.basename(part_path))
                    os.remove(part_path)

                os.remove(original_path)

        return send_file(zip_path, as_attachment=True)

    except Exception as e:
        return f"Erro no processamento: {str(e)}", 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
